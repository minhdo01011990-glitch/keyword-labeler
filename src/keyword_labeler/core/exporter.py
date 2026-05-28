from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise ImportError("openpyxl is required: pip install openpyxl") from e

try:
    import pandas as pd
except ImportError as e:
    raise ImportError("pandas is required: pip install pandas") from e

from .data_loader import KeywordRow

__all__ = ["export_excel", "export_markdown", "ExportResult"]

_INTENT_ORDER = ["informational", "commercial", "transactional", "navigational"]
_NO_GROUP_LABEL = "Chưa phân nhóm"

# Header fill colours per sheet (hex, no leading #)
_FILL_LABELED = "4472C4"   # blue
_FILL_REMOVED = "C0504D"   # red
_FILL_SUMMARY = "9BBB59"   # green


@dataclass
class ExportResult:
    output_path: Path
    labeled_count: int      # non-removed keywords exported
    removed_count: int      # removed keywords exported
    group_count: int        # distinct group names in Labeled sheet
    format: str             # "excel" or "markdown"


# ------------------------------------------------------------------ #
# Internal helpers
# ------------------------------------------------------------------ #

def _parse_group(group: str | None) -> tuple[str, str, str]:
    """Split 'chủ đề - hậu tố 1 - hậu tố 2' into (topic, suffix1, suffix2)."""
    if not group:
        return _NO_GROUP_LABEL, "", ""
    parts = [p.strip() for p in group.split(" - ", maxsplit=2)]
    topic = parts[0]
    suffix1 = parts[1] if len(parts) > 1 else ""
    suffix2 = parts[2] if len(parts) > 2 else ""
    return topic, suffix1, suffix2


def _build_labeled_df(keywords: list[KeywordRow]) -> pd.DataFrame:
    rows = []
    for kw in keywords:
        if kw.is_removed:
            continue
        topic, suffix1, suffix2 = _parse_group(kw.group)
        rows.append({
            "Keyword": kw.keyword,
            "Volume": kw.volume,
            "Nhóm": kw.group or _NO_GROUP_LABEL,
            "Chủ đề": topic,
            "Hậu tố 1": suffix1,
            "Hậu tố 2": suffix2,
            "Intent": kw.intent or "",
            "Thương hiệu": kw.brand_name or "",
        })
    if not rows:
        return pd.DataFrame(
            columns=["Keyword", "Volume", "Nhóm", "Chủ đề",
                     "Hậu tố 1", "Hậu tố 2", "Intent", "Thương hiệu"]
        )
    df = pd.DataFrame(rows)
    # Sort by group name, then volume descending within each group.
    return df.sort_values(["Nhóm", "Volume"], ascending=[True, False]).reset_index(drop=True)


def _build_removed_df(keywords: list[KeywordRow]) -> pd.DataFrame:
    rows = []
    for kw in keywords:
        if not kw.is_removed:
            continue
        rows.append({
            "Keyword": kw.keyword,
            "Volume": kw.volume,
            "Lý do lọc": kw.removal_reason or "",
        })
    if not rows:
        return pd.DataFrame(columns=["Keyword", "Volume", "Lý do lọc"])
    return pd.DataFrame(rows).sort_values("Volume", ascending=False).reset_index(drop=True)


def _build_summary_df(labeled_df: pd.DataFrame) -> pd.DataFrame:
    if labeled_df.empty:
        return pd.DataFrame(
            columns=["Nhóm", "Số keyword", "Tổng volume"]
            + [i.capitalize() for i in _INTENT_ORDER]
        )
    grouped = labeled_df.groupby("Nhóm", sort=False)
    rows = []
    for group_name, gdf in grouped:
        row: dict = {
            "Nhóm": group_name,
            "Số keyword": len(gdf),
            "Tổng volume": int(gdf["Volume"].sum()),
        }
        for intent in _INTENT_ORDER:
            row[intent.capitalize()] = int((gdf["Intent"] == intent).sum())
        rows.append(row)
    df = pd.DataFrame(rows)
    return df.sort_values("Số keyword", ascending=False).reset_index(drop=True)


def _df_to_markdown(df: pd.DataFrame) -> str:
    """Convert DataFrame to a GitHub-flavoured markdown table without tabulate."""
    if df.empty:
        return ""
    cols = list(df.columns)
    header = "| " + " | ".join(str(c) for c in cols) + " |"
    sep = "| " + " | ".join(["---"] * len(cols)) + " |"
    data_rows = [
        "| " + " | ".join(str(v) for v in row) + " |"
        for row in df.itertuples(index=False)
    ]
    return "\n".join([header, sep] + data_rows)


# ------------------------------------------------------------------ #
# Excel helpers
# ------------------------------------------------------------------ #

def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    if df.empty:
        return
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _style_header(ws, fill_hex: str) -> None:
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=False)


def _auto_width(ws) -> None:
    """Set column widths based on max content length (capped at 60 chars)."""
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


# ------------------------------------------------------------------ #
# Public API
# ------------------------------------------------------------------ #

def export_excel(
    keywords: list[KeywordRow],
    output_path: str | Path,
) -> ExportResult:
    """
    Export keyword labeling results to an Excel file with 3 sheets.

    Labeled  — all non-removed keywords: keyword, volume, group, topic, suffixes,
               intent, brand name. Sorted by group then volume descending.
    Removed  — filtered-out keywords with removal reason. Sorted by volume.
    Summary  — per-group stats: keyword count, total volume, intent breakdown.
               Sorted by keyword count descending.

    Args:
        keywords:    complete list of KeywordRow (including removed ones)
        output_path: destination .xlsx path; parent dirs created if needed
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    labeled_df = _build_labeled_df(keywords)
    removed_df = _build_removed_df(keywords)
    summary_df = _build_summary_df(labeled_df)

    wb = openpyxl.Workbook()

    ws_labeled = wb.active
    ws_labeled.title = "Labeled"
    _write_df_to_sheet(ws_labeled, labeled_df)
    if not labeled_df.empty:
        _style_header(ws_labeled, _FILL_LABELED)
        _auto_width(ws_labeled)

    ws_removed = wb.create_sheet("Removed")
    _write_df_to_sheet(ws_removed, removed_df)
    if not removed_df.empty:
        _style_header(ws_removed, _FILL_REMOVED)
        _auto_width(ws_removed)

    ws_summary = wb.create_sheet("Summary")
    _write_df_to_sheet(ws_summary, summary_df)
    if not summary_df.empty:
        _style_header(ws_summary, _FILL_SUMMARY)
        _auto_width(ws_summary)

    wb.save(output_path)

    return ExportResult(
        output_path=output_path,
        labeled_count=len(labeled_df),
        removed_count=len(removed_df),
        group_count=int(labeled_df["Nhóm"].nunique()) if not labeled_df.empty else 0,
        format="excel",
    )


def export_markdown(
    keywords: list[KeywordRow],
    output_path: str | Path,
    topic: str = "",
) -> ExportResult:
    """
    Export keyword labeling results to a Markdown file.

    Structure:
      # Title
      ## Tổng quan   (summary table: group / count / volume / intent breakdown)
      ## Chi tiết theo nhóm
        ### Group name (N kw, V vol)
        | Keyword | Volume | Intent | Thương hiệu |
        ...
      ## Keyword bị lọc

    Args:
        keywords:    complete list of KeywordRow (including removed ones)
        output_path: destination .md path; parent dirs created if needed
        topic:       SEO topic string used as document title
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    labeled_df = _build_labeled_df(keywords)
    removed_df = _build_removed_df(keywords)

    lines: list[str] = []

    title = f"Kết quả phân nhóm keyword{f': {topic}' if topic else ''}"
    lines.append(f"# {title}\n")

    if not labeled_df.empty:
        summary_df = _build_summary_df(labeled_df)
        lines.append("## Tổng quan\n")
        lines.append(_df_to_markdown(summary_df))
        lines.append("")

        lines.append("## Chi tiết theo nhóm\n")
        for group_name in summary_df["Nhóm"]:
            gdf = labeled_df[labeled_df["Nhóm"] == group_name][
                ["Keyword", "Volume", "Intent", "Thương hiệu"]
            ].reset_index(drop=True)
            total_vol = int(gdf["Volume"].sum())
            lines.append(f"### {group_name} ({len(gdf)} kw, {total_vol:,} vol)\n")
            lines.append(_df_to_markdown(gdf))
            lines.append("")

    if not removed_df.empty:
        lines.append(f"## Keyword bị lọc ({len(removed_df)})\n")
        lines.append(_df_to_markdown(removed_df))
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")

    return ExportResult(
        output_path=output_path,
        labeled_count=len(labeled_df),
        removed_count=len(removed_df),
        group_count=int(labeled_df["Nhóm"].nunique()) if not labeled_df.empty else 0,
        format="markdown",
    )
